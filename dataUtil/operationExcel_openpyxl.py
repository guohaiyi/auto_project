# -*- coding: UTF-8 -*-
from openpyxl import load_workbook
import os

proDir = os.path.split(os.path.realpath(__file__))[0]
excelPath = os.path.join(proDir, "Orc_dev_TestCase_test.xlsx")


class OperationExcel:

    def __init__(self, file_name=None):
        if file_name:
            self.file_name = file_name
        else:
            self.file_name = excelPath
        self.data = self.get_data()

    # 读取Excel表格
    def get_data(self):
        book = load_workbook(filename=self.file_name)
        tables = book.active
        return tables

    # 获取单元格的行数
    def get_lines(self):
        line = self.data.max_row
        return line

    # 通过单元格获取内容
    def get_cell_value(self, cell_name, row):
        value = self.data[cell_name + str(row)].value
        return value

    # 通过单元格坐标获取内容
    def get_coordinates_value(self, row, column):
        value = self.data.cell(row, column).value
        return value

    # 写入数据
    def write_result(self, write_name, row, write_value):
        wb = load_workbook(filename=self.file_name)
        ws = wb.active
        ws[write_name + str(row)] = write_value
        wb.save(filename=self.file_name)


if __name__ == "__main__":
    s = OperationExcel()
    ss = s.get_coordinates_value(1, 2)
    print(ss)
